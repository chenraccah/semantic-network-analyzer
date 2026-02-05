"""
Export service for analysis data in multiple formats.
"""
import io
import json
import networkx as nx
import openpyxl
from typing import List, Dict, Any


def export_excel(nodes: List[Dict], edges: List[Dict], stats: Dict, group_names: List[str], group_keys: List[str]) -> bytes:
    """Export to Excel with multiple sheets: Words, Edges, Stats."""
    wb = openpyxl.Workbook()
    
    # Words sheet
    ws_words = wb.active
    ws_words.title = "Words"
    # Headers
    headers = ["Word"]
    for i, name in enumerate(group_names):
        headers.extend([f"{name}_Count", f"{name}_Score"])
    headers.append("Avg_Score")
    if len(group_names) == 2:
        headers.extend(["Difference", "Emphasis"])
    for i, name in enumerate(group_names):
        headers.extend([f"{name}_Cluster", f"{name}_Betweenness"])
    ws_words.append(headers)
    
    for node in nodes:
        row = [node.get("word", "")]
        for key in group_keys:
            row.append(node.get(f"{key}_count", 0))
            row.append(node.get(f"{key}_normalized", 0))
        row.append(node.get("avg_normalized", 0))
        if len(group_names) == 2:
            diff = node.get("difference", 0)
            emphasis = "Balanced" if abs(diff) < 10 else (group_names[0] if diff > 0 else group_names[1])
            row.extend([diff, emphasis])
        for key in group_keys:
            row.append(node.get(f"{key}_cluster", -1))
            row.append(node.get(f"{key}_betweenness", 0))
        ws_words.append(row)
    
    # Edges sheet
    ws_edges = wb.create_sheet("Edges")
    ws_edges.append(["From", "To", "Weight", "Semantic_Similarity", "Edge_Type"])
    for edge in edges:
        ws_edges.append([
            edge.get("from", ""),
            edge.get("to", ""),
            edge.get("weight", 0),
            edge.get("semantic_similarity", ""),
            edge.get("edge_type", "cooccurrence"),
        ])
    
    # Stats sheet
    ws_stats = wb.create_sheet("Stats")
    ws_stats.append(["Metric", "Value"])
    for key, value in stats.items():
        ws_stats.append([str(key), str(value)])
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_json(nodes: List[Dict], edges: List[Dict], stats: Dict, group_names: List[str], group_keys: List[str]) -> bytes:
    """Export to JSON."""
    data = {
        "nodes": nodes,
        "edges": edges,
        "stats": stats,
        "group_names": group_names,
        "group_keys": group_keys,
    }
    return json.dumps(data, indent=2, default=str).encode("utf-8")


def _build_nx_graph(nodes: List[Dict], edges: List[Dict], group_keys: List[str]) -> nx.Graph:
    """Build a NetworkX graph from node and edge data."""
    G = nx.Graph()
    for node in nodes:
        attrs = {k: v for k, v in node.items() if k != "word" and v is not None}
        G.add_node(node["word"], **attrs)
    for edge in edges:
        attrs = {k: v for k, v in edge.items() if k not in ("from", "to") and v is not None}
        G.add_edge(edge["from"], edge["to"], **attrs)
    return G


def export_graphml(nodes: List[Dict], edges: List[Dict], stats: Dict, group_names: List[str], group_keys: List[str]) -> bytes:
    """Export to GraphML format."""
    G = _build_nx_graph(nodes, edges, group_keys)
    buf = io.BytesIO()
    nx.write_graphml(G, buf)
    buf.seek(0)
    return buf.getvalue()


def export_gexf(nodes: List[Dict], edges: List[Dict], stats: Dict, group_names: List[str], group_keys: List[str]) -> bytes:
    """Export to GEXF format."""
    G = _build_nx_graph(nodes, edges, group_keys)
    buf = io.BytesIO()
    nx.write_gexf(G, buf)
    buf.seek(0)
    return buf.getvalue()
